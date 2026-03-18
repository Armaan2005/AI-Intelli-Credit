"""
pdf_parser.py
-------------
Upgrades your existing extract_text() with:
  - Document type auto-detection (GST/Bank/Annual Report/Legal)
  - Page-level extraction with metadata
  - Better OCR fallback threshold
  - File format validation
  - Excel/CSV support (for GST returns)

DROP-IN REPLACEMENT — same function name extract_text(), richer output.
"""
import os
from typing import Union

try:
    import fitz  # PyMuPDF
    FITZ_AVAILABLE = True
except ImportError:
    FITZ_AVAILABLE = False

from app.services.extraction.ocr_pipeline import ocr_extract, get_text


# ─────────────────────────────────────────────────────────────────────────────
#  Document type detection keywords
# ─────────────────────────────────────────────────────────────────────────────
DOC_TYPE_SIGNALS = {
    "gst_return":      ["GSTIN", "GSTR", "GST Return", "ITC", "outward supplies", "B2B"],
    "bank_statement":  ["Account Number", "IFSC", "Closing Balance", "Debit", "Credit", "NEFT", "RTGS"],
    "itr":             ["Income Tax Return", "PAN", "Assessment Year", "TDS", "ITR-"],
    "annual_report":   ["Directors' Report", "Auditor", "Balance Sheet", "Standalone", "Consolidated"],
    "legal_notice":    ["NCLT", "DRT", "Court", "Petitioner", "Respondent", "Legal Notice", "Advocate"],
    "sanction_letter": ["Sanction Letter", "Sanctioned Limit", "Rate of Interest", "Security", "Terms and Conditions"],
    "rating_report":   ["Rating", "CRISIL", "ICRA", "CARE", "Outlook", "Upgrade", "Downgrade"],
}


def _detect_doc_type(text: str) -> str:
    """Auto-detect document type from text content."""
    text_upper = text[:3000].upper()
    scores = {}
    for doc_type, keywords in DOC_TYPE_SIGNALS.items():
        score = sum(1 for kw in keywords if kw.upper() in text_upper)
        if score > 0:
            scores[doc_type] = score
    if not scores:
        return "unknown"
    return max(scores, key=scores.get)


def _extract_excel_text(file_path: str) -> str:
    """Extract text from Excel files (GST returns often come as .xlsx)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text_parts.append(f"[Sheet: {sheet}]")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        return f"Excel extraction failed: {e}"


def _extract_csv_text(file_path: str) -> str:
    """Extract text from CSV files."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        return f"CSV read failed: {e}"


def extract_text(file_path: str) -> Union[str, dict]:
    """
    UPGRADED version of your extract_text().

    Improvements over original:
    - Supports PDF, Excel (.xlsx), CSV file types
    - Document type auto-detection
    - OCR threshold raised to 100 chars (your 50 was too low)
    - Page-level extraction with metadata
    - Structured return with doc_type, page_count, ocr_used

    Backward-compatible: result is a string OR dict with "text" key.
    Use get_plain_text(result) helper to always get string.
    """
    if not os.path.exists(file_path):
        return {"text": "", "error": f"File not found: {file_path}", "doc_type": "unknown"}

    ext = os.path.splitext(file_path)[1].lower()

    # ── Excel files ───────────────────────────────────────────────────────────
    if ext in (".xlsx", ".xls", ".xlsm"):
        text = _extract_excel_text(file_path)
        return {
            "text":       text,
            "doc_type":   _detect_doc_type(text),
            "page_count": 1,
            "ocr_used":   False,
            "file_type":  "excel",
        }

    # ── CSV files ────────────────────────────────────────────────────────────
    if ext == ".csv":
        text = _extract_csv_text(file_path)
        return {
            "text":       text,
            "doc_type":   _detect_doc_type(text),
            "page_count": 1,
            "ocr_used":   False,
            "file_type":  "csv",
        }

    # ── PDF files ─────────────────────────────────────────────────────────────
    if not FITZ_AVAILABLE:
        ocr_result = ocr_extract(file_path)
        text = get_text(ocr_result)
        return {
            "text":       text,
            "doc_type":   _detect_doc_type(text),
            "page_count": ocr_result.get("total_pages", 0) if isinstance(ocr_result, dict) else 0,
            "ocr_used":   True,
            "file_type":  "pdf",
        }

    try:
        doc       = fitz.open(file_path)
        full_text = ""
        pages     = []

        for page in doc:
            page_text = page.get_text().strip()
            pages.append({
                "page":       page.number + 1,
                "char_count": len(page_text),
                "text":       page_text,
            })
            full_text += page_text + "\n\n"

        doc.close()

        # ── OCR fallback: if less than 100 chars extracted (your original was 50)
        # 50 chars is too low — a page with just a header would pass
        if len(full_text.strip()) < 100:
            print(f"📄 PDF has <100 chars extracted — switching to OCR for {file_path}")
            ocr_result = ocr_extract(file_path)
            ocr_text   = get_text(ocr_result)
            return {
                "text":       ocr_text,
                "doc_type":   _detect_doc_type(ocr_text),
                "page_count": len(pages),
                "ocr_used":   True,
                "file_type":  "pdf",
                "ocr_confidence": ocr_result.get("avg_confidence", 0) if isinstance(ocr_result, dict) else 0,
            }

        return {
            "text":       full_text.strip(),
            "doc_type":   _detect_doc_type(full_text),
            "page_count": len(pages),
            "pages":      pages,
            "ocr_used":   False,
            "file_type":  "pdf",
        }

    except Exception as e:
        print(f"⚠️ PyMuPDF failed: {e} — trying OCR")
        ocr_result = ocr_extract(file_path)
        text       = get_text(ocr_result)
        return {
            "text":     text,
            "doc_type": _detect_doc_type(text),
            "ocr_used": True,
            "error":    str(e),
            "file_type": "pdf",
        }


def get_plain_text(result: Union[str, dict]) -> str:
    """Helper: always returns plain string from extract_text() output."""
    if isinstance(result, str):
        return result
    return result.get("text", "")